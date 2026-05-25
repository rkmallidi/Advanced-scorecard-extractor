"""
Tkinter UI for the JEE Advanced response extractor.

Run from the project root:
python src/jee_response_extractor_ui.py
"""

from pathlib import Path
import os
import queue
import threading
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk

from openpyxl.styles import Font
from playwright.sync_api import sync_playwright


SCRIPT_DIR = Path(__file__).resolve().parent
os.chdir(SCRIPT_DIR)

import jee_response_extractor_With_Key as extractor


class ExtractorApp:

    def __init__(self, root):
        self.root = root
        self.root.title("JEE Advanced Scorecard Extractor")
        self.root.geometry("1100x680")
        self.root.minsize(900, 560)

        self.events = queue.Queue()
        self.candidates = []
        self.candidate_summaries = {}
        self.running = False
        self.stop_requested = False
        self.output_mode = extractor.CREATE_NEW_OUTPUT

        self.total_var = tk.StringVar(value="Total: 0")
        self.success_var = tk.StringVar(value="Saved: 0")
        self.failed_var = tk.StringVar(value="Failed: 0")
        self.current_var = tk.StringVar(value="Current: -")
        self.selected_candidate_var = tk.StringVar(
            value="Select a saved candidate to preview the score summary."
        )
        self.run_headless = extractor.HEADLESS

        self._build_ui()
        self._load_candidates()
        self._poll_events()

    def _build_ui(self):
        root = self.root
        style = ttk.Style(root)
        style.configure(
            "Summary.TLabel",
            font=("Segoe UI", 10, "bold")
        )
        style.configure(
            "MetricValue.TLabel",
            font=("Segoe UI", 18, "bold")
        )
        style.configure(
            "MetricLabel.TLabel",
            font=("Segoe UI", 9, "bold")
        )

        toolbar = ttk.Frame(root, padding=(12, 10))
        toolbar.pack(fill=tk.X)

        self.start_button = ttk.Button(
            toolbar,
            text="Start Extraction",
            command=self.start_extraction
        )
        self.start_button.pack(side=tk.LEFT)

        self.reload_button = ttk.Button(
            toolbar,
            text="Reload Candidates",
            command=self._load_candidates
        )
        self.reload_button.pack(side=tk.LEFT, padx=(8, 0))

        self.open_output_button = ttk.Button(
            toolbar,
            text="Open Output File",
            command=self._open_output_file,
            state=(
                tk.NORMAL
                if extractor.OUTPUT_FILE.exists()
                else tk.DISABLED
            )
        )
        self.open_output_button.pack(side=tk.LEFT, padx=(8, 0))

        stats = ttk.Frame(root, padding=(12, 0, 12, 10))
        stats.pack(fill=tk.X)

        metric_cards = [
            ("Candidates", self.total_var),
            ("Saved", self.success_var),
            ("Failed", self.failed_var)
        ]

        for label, variable in metric_cards:
            card = ttk.Frame(
                stats,
                padding=(12, 8),
                relief=tk.GROOVE
            )
            card.pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)

            ttk.Label(
                card,
                text=label.upper(),
                style="MetricLabel.TLabel"
            ).pack(anchor=tk.W)

            ttk.Label(
                card,
                textvariable=variable,
                style="MetricValue.TLabel"
            ).pack(anchor=tk.W)

        ttk.Label(
            root,
            textvariable=self.current_var,
            padding=(12, 0, 12, 8),
            style="Summary.TLabel"
        ).pack(fill=tk.X)

        self.progress = ttk.Progressbar(
            root,
            mode="determinate"
        )
        self.progress.pack(fill=tk.X, padx=12, pady=(0, 10))

        main_pane = ttk.PanedWindow(root, orient=tk.HORIZONTAL)
        main_pane.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 0))

        table_frame = ttk.Frame(main_pane)
        main_pane.add(table_frame, weight=4)

        columns = (
            "number",
            "username",
            "dob",
            "status",
            "marks",
            "attempted",
            "correct",
            "partial",
            "wrong",
            "message"
        )

        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            height=18
        )

        self.tree.heading("number", text="#")
        self.tree.heading("username", text="Username")
        self.tree.heading("dob", text="DOB")
        self.tree.heading("status", text="Status")
        self.tree.heading("marks", text="Marks")
        self.tree.heading("attempted", text="Attempted")
        self.tree.heading("correct", text="Correct")
        self.tree.heading("partial", text="Partial")
        self.tree.heading("wrong", text="Wrong")
        self.tree.heading("message", text="Message")

        self.tree.column("number", width=50, stretch=False, anchor=tk.CENTER)
        self.tree.column("username", width=150, stretch=False)
        self.tree.column("dob", width=105, stretch=False)
        self.tree.column("status", width=105, stretch=False, anchor=tk.CENTER)
        self.tree.column("marks", width=80, stretch=False, anchor=tk.CENTER)
        self.tree.column("attempted", width=90, stretch=False, anchor=tk.CENTER)
        self.tree.column("correct", width=75, stretch=False, anchor=tk.CENTER)
        self.tree.column("partial", width=75, stretch=False, anchor=tk.CENTER)
        self.tree.column("wrong", width=75, stretch=False, anchor=tk.CENTER)
        self.tree.column("message", width=300, stretch=True)

        self.tree.tag_configure("Saved", background="#E7F6EC")
        self.tree.tag_configure("Failed", background="#FDECEC")
        self.tree.tag_configure("Processing", background="#FFF4D6")

        scrollbar = ttk.Scrollbar(
            table_frame,
            orient=tk.VERTICAL,
            command=self.tree.yview
        )
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.bind("<<TreeviewSelect>>", self._on_candidate_selected)

        summary_frame = ttk.Frame(main_pane, padding=(12, 0, 0, 0))
        main_pane.add(summary_frame, weight=2)

        ttk.Label(
            summary_frame,
            text="Selected Candidate Summary",
            style="Summary.TLabel"
        ).pack(anchor=tk.W, pady=(0, 6))

        ttk.Label(
            summary_frame,
            textvariable=self.selected_candidate_var,
            wraplength=330
        ).pack(anchor=tk.W, pady=(0, 10))

        self.subject_tree = ttk.Treeview(
            summary_frame,
            columns=(
                "subject",
                "paper1",
                "paper2",
                "total",
                "correct",
                "partial",
                "wrong"
            ),
            show="headings",
            height=5
        )

        for column, heading, width in (
            ("subject", "Subject", 95),
            ("paper1", "P1", 48),
            ("paper2", "P2", 48),
            ("total", "Total", 58),
            ("correct", "C", 42),
            ("partial", "P", 42),
            ("wrong", "W", 42)
        ):
            self.subject_tree.heading(column, text=heading)
            self.subject_tree.column(
                column,
                width=width,
                stretch=False,
                anchor=tk.CENTER
            )

        self.subject_tree.pack(fill=tk.X)

        log_frame = ttk.Frame(root, padding=12)
        log_frame.pack(fill=tk.BOTH)

        ttk.Label(log_frame, text="Activity").pack(anchor=tk.W)

        self.log = tk.Text(
            log_frame,
            height=7,
            wrap=tk.WORD,
            state=tk.DISABLED
        )
        self.log.pack(fill=tk.BOTH, expand=False)

    def _load_candidates(self):
        if self.running:
            return

        self.tree.delete(*self.tree.get_children())
        self.subject_tree.delete(*self.subject_tree.get_children())
        self.candidate_summaries = {}
        self.candidates = []

        try:
            self.candidates = extractor.read_candidate_excel()

        except Exception as exc:
            messagebox.showerror(
                "Could not load candidates",
                str(exc)
            )
            self._set_counts(0, 0, 0)
            return

        for index, candidate in enumerate(self.candidates, 1):
            self.tree.insert(
                "",
                tk.END,
                iid=str(index - 1),
                values=(
                    index,
                    extractor.clean_text(candidate.get("USERNAME", "")),
                    extractor.clean_text(candidate.get("DOB", "")),
                    "Pending",
                    "",
                    "",
                    "",
                    "",
                    "",
                    ""
                )
            )

        self.progress.configure(
            maximum=max(len(self.candidates), 1),
            value=0
        )
        self._set_counts(len(self.candidates), 0, 0)
        self.current_var.set("Current: -")
        self.selected_candidate_var.set(
            "Select a saved candidate to preview the score summary."
        )
        self._append_log(f"Loaded {len(self.candidates)} candidates.")

    def start_extraction(self):
        if self.running:
            return

        if not self.candidates:
            messagebox.showwarning(
                "No candidates",
                "No candidates are available to process."
            )
            return

        if extractor.OUTPUT_FILE.exists():
            append_existing = messagebox.askyesnocancel(
                "Output file",
                "Append responses to the existing output file?\n\n"
                "Yes: append new responses to the current workbook.\n"
                "No: create a new workbook from scratch.\n"
                "Cancel: do not start extraction."
            )

            if append_existing is None:
                return

            self.output_mode = (
                extractor.APPEND_OUTPUT
                if append_existing
                else extractor.CREATE_NEW_OUTPUT
            )

        else:
            self.output_mode = extractor.CREATE_NEW_OUTPUT

        self.running = True
        self.run_headless = extractor.HEADLESS
        self.stop_requested = False
        self.start_button.configure(state=tk.DISABLED)
        self.reload_button.configure(state=tk.DISABLED)
        self.open_output_button.configure(state=tk.DISABLED)
        self.candidate_summaries = {}
        self.subject_tree.delete(*self.subject_tree.get_children())
        self.selected_candidate_var.set(
            "Summary will appear here as candidates are saved."
        )

        for item in self.tree.get_children():
            values = list(self.tree.item(item, "values"))
            values[3] = "Pending"
            values[4] = ""
            values[5] = ""
            values[6] = ""
            values[7] = ""
            values[8] = ""
            values[9] = ""
            self.tree.item(item, values=values, tags=())

        self.progress.configure(value=0)

        browser_mode = (
            "headless"
            if self.run_headless
            else "visible"
        )
        self._append_log(
            f"Starting extraction with {browser_mode} browser mode."
        )

        if self.output_mode == extractor.APPEND_OUTPUT:
            self._append_log("Responses will be appended to the existing output file.")
        else:
            self._append_log("A new output file will be created.")

        worker = threading.Thread(
            target=self._run_extraction,
            daemon=True
        )
        worker.start()

    def _run_extraction(self):
        successful = 0
        failed_candidates = []

        try:
            answer_key_map = extractor.read_answer_key()

            workbook, appended_to_existing = extractor.create_output_workbook(
                self.output_mode
            )

            if appended_to_existing:
                self.events.put(
                    (
                        "log",
                        f"Loaded existing workbook: {extractor.OUTPUT_FILE}"
                    )
                )

            with sync_playwright() as p:
                browser = p.chromium.launch(
                    headless=self.run_headless,
                    slow_mo=0 if self.run_headless else extractor.SLOW_MO
                )

                for index, candidate in enumerate(self.candidates):
                    username = extractor.clean_text(
                        candidate.get("USERNAME", "")
                    )
                    self.events.put(
                        ("processing", index, username)
                    )

                    context = browser.new_context()

                    try:
                        candidate_info, results, error_message = (
                            extractor.process_candidate(
                                context,
                                candidate,
                                answer_key_map
                            )
                        )

                        if candidate_info and results:
                            candidate_id = extractor.clean_text(
                                candidate_info.get("Candidate ID", "")
                            )

                            if extractor.candidate_exists_in_workbook(
                                workbook,
                                candidate_info
                            ):
                                workbook.save(extractor.OUTPUT_FILE)
                                self.events.put(
                                    (
                                        "skipped",
                                        index,
                                        f"Skipped {candidate_id}: candidate already exists in output file"
                                    )
                                )
                                continue

                            summary = extractor.calculate_candidate_summary(
                                results
                            )
                            detail_sheet_name = extractor.write_candidate_detail_sheet(
                                workbook,
                                candidate_info,
                                results
                            )
                            extractor.write_candidate_sheet(
                                workbook,
                                candidate_info,
                                results,
                                detail_sheet_name
                            )
                            workbook.save(extractor.OUTPUT_FILE)
                            successful += 1
                            self.events.put(
                                (
                                    "saved",
                                    index,
                                    f"Saved row and detail sheet for {username}",
                                    candidate_info,
                                    summary
                                )
                            )

                        else:
                            failed = extractor.build_failed_candidate_record(
                                candidate,
                                "Failed",
                                error_message or "No candidate details / question data extracted"
                            )
                            failed_candidates.append(failed)
                            extractor.write_failed_candidates_excel(
                                failed_candidates
                            )
                            self.events.put(
                                (
                                    "failed",
                                    index,
                                    failed["Error"]
                                )
                            )

                    except Exception as exc:
                        failed = extractor.build_failed_candidate_record(
                            candidate,
                            "Failed",
                            str(exc)
                        )
                        failed_candidates.append(failed)
                        extractor.write_failed_candidates_excel(
                            failed_candidates
                        )
                        self.events.put(
                            (
                                "failed",
                                index,
                                str(exc)
                            )
                        )

                    finally:
                        context.close()

                browser.close()

            if not workbook.sheetnames:
                ws = workbook.create_sheet(
                    title="No Successful Candidates"
                )
                ws["A1"] = "No successful candidates processed"
                ws["A1"].font = Font(bold=True)

            workbook.save(extractor.OUTPUT_FILE)
            extractor.write_failed_candidates_excel(failed_candidates)

            self.events.put(
                (
                    "done",
                    successful,
                    len(failed_candidates)
                )
            )

        except Exception as exc:
            self.events.put(
                (
                    "fatal",
                    str(exc)
                )
            )

    def _poll_events(self):
        try:
            while True:
                event = self.events.get_nowait()
                self._handle_event(event)

        except queue.Empty:
            pass

        self.root.after(200, self._poll_events)

    def _handle_event(self, event):
        event_type = event[0]

        if event_type == "processing":
            _, index, username = event
            self.current_var.set(f"Current: {username}")
            self._update_row(index, "Processing", "Logging in and extracting")
            self._append_log(f"Processing {username}")
            return

        if event_type == "log":
            _, message = event
            self._append_log(message)
            return

        if event_type == "saved":
            _, index, message, candidate_info, summary = event
            self.candidate_summaries[index] = {
                "candidate_info": candidate_info,
                "summary": summary
            }
            self._update_saved_row(
                index,
                message,
                summary
            )
            self._increment_progress()
            self._refresh_counts()
            self._render_candidate_summary(index)
            self._append_log(message)
            return

        if event_type == "failed":
            _, index, message = event
            self._update_row(index, "Failed", message)
            self._increment_progress()
            self._refresh_counts()
            self._append_log(f"Failed: {message}")
            return

        if event_type == "skipped":
            _, index, message = event
            self._update_row(index, "Skipped", message)
            self._increment_progress()
            self._refresh_counts()
            self._append_log(message)
            return

        if event_type == "done":
            _, successful, failed = event
            self.running = False
            self.start_button.configure(state=tk.NORMAL)
            self.reload_button.configure(state=tk.NORMAL)
            self.open_output_button.configure(state=tk.NORMAL)
            self.current_var.set("Current: -")
            self._append_log(
                f"Completed. Saved: {successful}, Failed: {failed}"
            )
            messagebox.showinfo(
                "Extraction complete",
                f"Saved: {successful}\nFailed: {failed}"
            )
            return

        if event_type == "fatal":
            _, message = event
            self.running = False
            self.start_button.configure(state=tk.NORMAL)
            self.reload_button.configure(state=tk.NORMAL)
            self.open_output_button.configure(
                state=(
                    tk.NORMAL
                    if extractor.OUTPUT_FILE.exists()
                    else tk.DISABLED
                )
            )
            self._append_log(f"Fatal error: {message}")
            messagebox.showerror(
                "Extraction stopped",
                message
            )

    def _open_output_file(self):
        output_file = extractor.OUTPUT_FILE

        if not output_file.exists():
            messagebox.showwarning(
                "Output file not found",
                f"Could not find:\n{output_file}"
            )
            return

        try:
            os.startfile(output_file)

        except Exception as exc:
            messagebox.showerror(
                "Could not open output file",
                str(exc)
            )

    def _update_row(self, index, status, message):
        item_id = str(index)

        if not self.tree.exists(item_id):
            return

        values = list(self.tree.item(item_id, "values"))
        values[3] = status
        values[9] = message
        self.tree.item(item_id, values=values, tags=(status,))
        self.tree.see(item_id)

    def _update_saved_row(self, index, message, summary):
        item_id = str(index)

        if not self.tree.exists(item_id):
            return

        result_counts = {
            "Correct": 0,
            "Partial": 0,
            "Wrong": 0
        }

        for counts in summary["subject_result_counts"].values():
            for result in result_counts:
                result_counts[result] += counts.get(result, 0)

        values = list(self.tree.item(item_id, "values"))
        values[3] = "Saved"
        values[4] = summary["total_marks"]
        values[5] = (
            f'{summary["attempted"]}/'
            f'{summary["total_questions"]}'
        )
        values[6] = result_counts["Correct"]
        values[7] = result_counts["Partial"]
        values[8] = result_counts["Wrong"]
        values[9] = message
        self.tree.item(item_id, values=values, tags=("Saved",))
        self.tree.see(item_id)

    def _increment_progress(self):
        self.progress.configure(
            value=self.progress["value"] + 1
        )

    def _refresh_counts(self):
        saved = 0
        failed = 0

        for item in self.tree.get_children():
            status = self.tree.item(item, "values")[3]

            if status == "Saved":
                saved += 1

            elif status == "Failed":
                failed += 1

        self._set_counts(
            len(self.candidates),
            saved,
            failed
        )

    def _set_counts(
        self,
        total,
        successful,
        failed
    ):
        self.total_var.set(str(total))
        self.success_var.set(str(successful))
        self.failed_var.set(str(failed))

    def _on_candidate_selected(self, _event=None):
        selected = self.tree.selection()

        if not selected:
            return

        self._render_candidate_summary(
            int(selected[0])
        )

    def _render_candidate_summary(self, index):
        self.subject_tree.delete(*self.subject_tree.get_children())

        summary_info = self.candidate_summaries.get(index)

        if not summary_info:
            self.selected_candidate_var.set(
                "Summary is available after this candidate is saved."
            )
            return

        candidate_info = summary_info["candidate_info"]
        summary = summary_info["summary"]

        name = extractor.clean_text(
            candidate_info.get("Candidate Name", "")
        )
        candidate_id = extractor.clean_text(
            candidate_info.get("Candidate ID", "")
        )

        label = name or candidate_id or f"Candidate {index + 1}"

        self.selected_candidate_var.set(
            f"{label}\n"
            f"Total Marks: {summary['total_marks']} | "
            f"Attempted: {summary['attempted']}/"
            f"{summary['total_questions']}"
        )

        for subject in (
            "Mathematics",
            "Physics",
            "Chemistry"
        ):
            marks = summary["subject_paper_marks"].get(
                subject,
                {}
            )
            counts = summary["subject_result_counts"].get(
                subject,
                {}
            )
            paper_1 = marks.get("Paper 1", 0)
            paper_2 = marks.get("Paper 2", 0)

            self.subject_tree.insert(
                "",
                tk.END,
                values=(
                    subject.replace("Mathematics", "Math"),
                    paper_1,
                    paper_2,
                    paper_1 + paper_2,
                    counts.get("Correct", 0),
                    counts.get("Partial", 0),
                    counts.get("Wrong", 0)
                )
            )

    def _append_log(self, message):
        self.log.configure(state=tk.NORMAL)
        self.log.insert(tk.END, f"{message}\n")
        self.log.see(tk.END)
        self.log.configure(state=tk.DISABLED)


def main():
    root = tk.Tk()
    app = ExtractorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
