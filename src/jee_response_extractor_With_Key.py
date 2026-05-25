"""
JEE Advanced Multi-Candidate Response Extractor
================================================

FEATURES
--------
1. Reads multiple candidates from Excel
2. Logs into JEE Advanced portal
3. Opens response sheets automatically
4. Extracts:
    - Candidate Details
    - Question Details
    - Question Type
    - Status
    - Chosen Option / Given Answer
    - Internal Option Mapping
    - Internal Answer IDs
    - Correct Key Mapping
5. Supports:
    - MCQ
    - MSQ
    - SA / NAT
6. Creates:
    - One Excel sheet with one row per candidate
7. Stable DOM parsing using BeautifulSoup
8. Handles malformed HTML correctly
9. Stores Option Mapping as JSON-like string

INSTALL
-------
pip install playwright pandas openpyxl beautifulsoup4

FIRST TIME
----------
playwright install

INPUT FILES
-----------

1. input/candidates.xlsx

| USERNAME     | PASSWORD   | DOB        |
|--------------|------------|------------|
| R72121106620 | 9493205542 | 10-01-2008 |

2. jeekey/jee_answer_key.xlsx

Sheet Name:
JEE Responses

Required Columns:
- Question id
- Correct answer

RUN
---
python jee_response_extractor.py
"""

from playwright.sync_api import sync_playwright
from urllib.parse import urljoin

from bs4 import BeautifulSoup

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pandas as pd
import traceback
import re

# ============================================================
# CONFIGURATION
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parents[1]

INPUT_EXCEL = PROJECT_ROOT / "input" / "candidates.xlsx"

ANSWER_KEY_FILE = PROJECT_ROOT / "JeeKey" / "ADV 2026 Key & Marking Scheme.xlsx"
LOGIN_URL = "https://cportal.jeeadv.ac.in/"
#LOGIN_URL = "https://candidate-portal.jeeadv.ac.in/authenticate"

DISPLAY_RESPONSE_URL = (
    "https://cportal.jeeadv.ac.in/display-response"
)

OUTPUT_FILE = PROJECT_ROOT / "output" / "jee_all_candidates_output.xlsx"

FAILED_OUTPUT_FILE = PROJECT_ROOT / "output" / "failed_candidates.xlsx"

HEADLESS = False
SLOW_MO = 300

CREATE_NEW_OUTPUT = "new"
APPEND_OUTPUT = "append"

# ============================================================
# COMMON HELPERS
# ============================================================

def clean_text(text):

    if text is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(text)
    ).strip()


def create_output_workbook(output_mode=CREATE_NEW_OUTPUT):

    if output_mode == APPEND_OUTPUT and OUTPUT_FILE.exists():
        workbook = load_workbook(OUTPUT_FILE)
        return workbook, True

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    return workbook, False


def ask_output_mode():

    if not OUTPUT_FILE.exists():
        print(
            f"No existing output file found. A new file will be created:\n{OUTPUT_FILE}"
        )
        return CREATE_NEW_OUTPUT

    while True:
        choice = input(
            "\nAppend responses to the existing output file? "
            "Enter Y to append or N to create a new file: "
        ).strip().lower()

        if choice in ("y", "yes"):
            return APPEND_OUTPUT

        if choice in ("n", "no"):
            return CREATE_NEW_OUTPUT

        print("Please enter Y or N.")

def parse_mark(value):

    text = clean_text(value)

    if not text:
        return 0

    try:
        mark = Decimal(text)

    except InvalidOperation:
        return 0

    if mark == mark.to_integral_value():
        return int(mark)

    return float(mark)

# ============================================================
# DOB FORMATTER
# ============================================================

def format_dob(dob):

    dob = str(dob).strip()

    if " " in dob:
        dob = dob.split(" ")[0]

    formats = [
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%Y/%m/%d"
    ]

    for fmt in formats:

        try:

            return datetime.strptime(
                dob,
                fmt
            ).strftime("%d-%m-%Y")

        except:
            pass

    raise Exception(
        f"Unsupported DOB format: {dob}"
    )

# ============================================================
# READ CANDIDATE EXCEL
# ============================================================

def read_candidate_excel():

    df = pd.read_excel(
        INPUT_EXCEL,
        dtype=str
    )

    required_cols = [
        "USERNAME",
        "PASSWORD",
        "DOB"
    ]

    for col in required_cols:

        if col not in df.columns:

            raise Exception(
                f"Missing column: {col}"
            )

    return df.to_dict(
        orient="records"
    )

# ============================================================
# READ ANSWER KEY
# ============================================================

def read_answer_key():

    df = pd.read_excel(
        ANSWER_KEY_FILE,        
        dtype=str
    )

    df.columns = [
        str(col).strip()
        for col in df.columns
    ]

    question_col = None
    key_col = None
    positive_mark_col = None
    negative_mark_col = None
    paper_col = None
    subject_col = None
    section_col = None
    question_number_col = None

    for col in df.columns:

        lower_col = col.lower()

        if lower_col == "paper":
            paper_col = col

        if lower_col == "subject":
            subject_col = col

        if lower_col == "sec":
            section_col = col

        if lower_col == "q":
            question_number_col = col

        if (
            "question" in lower_col
            and "id" in lower_col
        ):
            question_col = col

        if lower_col == "key":
            key_col = col

        if lower_col in ("+mark", "+marks", "positive mark", "positive marks"):
            positive_mark_col = col

        if lower_col in ("-ve mark", "-vemark", "-ve marks", "negative mark", "negative marks"):
            negative_mark_col = col

    if not question_col or not key_col:

        raise Exception(
            "Question ID / Key columns missing"
        )

    answer_map = {}

    for _, row in df.iterrows():

        qid = clean_text(
            row.get(question_col, "")
        )

        key = clean_text(
            row.get(key_col, "")
        )

        positive_mark = parse_mark(
            row.get(positive_mark_col, "")
            if positive_mark_col
            else 0
        )

        negative_mark = parse_mark(
            row.get(negative_mark_col, "")
            if negative_mark_col
            else 0
        )

        if qid:
            answer_map[qid] = {
                "Paper": clean_text(
                    row.get(paper_col, "")
                    if paper_col
                    else ""
                ),
                "Subject": clean_text(
                    row.get(subject_col, "")
                    if subject_col
                    else ""
                ),
                "Sec": clean_text(
                    row.get(section_col, "")
                    if section_col
                    else ""
                ),
                "Q": clean_text(
                    row.get(question_number_col, "")
                    if question_number_col
                    else ""
                ),
                "Key": key,
                "+mark": positive_mark,
                "-ve mark": negative_mark
            }

    print(
        f"Loaded {len(answer_map)} answer keys"
    )

    return answer_map

# ============================================================
# CANDIDATE INFO EXTRACTION
# ============================================================

def extract_candidate_info(page):

    body_text = page.locator(
        "body"
    ).inner_text()

    candidate_info = {}

    fields = [
        "Candidate ID",
        "Candidate Name",
        "Test Center Name",
        "Test Date",
        "Test Time",
        "Subject"
    ]

    for field in fields:

        pattern = rf"{field}\s*(.+)"

        match = re.search(
            pattern,
            body_text
        )

        if match:

            candidate_info[field] = clean_text(
                match.group(1)
            )

        else:

            candidate_info[field] = ""

    return candidate_info

# ============================================================
# OPTION HELPERS
# ============================================================

def extract_internal_option_id(img_tag):

    """
    Example:
    q4o3.png -> 3
    """

    if not img_tag:
        return ""

    name = img_tag.get("name", "")

    match = re.search(
        r"o(\d+)\.png",
        name
    )

    if match:
        return match.group(1)

    return ""

def internal_option_to_letter(value):

    return {
        "1": "A",
        "2": "B",
        "3": "C",
        "4": "D"
    }.get(str(value), "")

def clean_option_text(value):

    if not value:
        return ""

    return (
        str(value)
        .replace("--", "")
        .strip()
    )

def option_mapping_to_text(option_mapping):

    return " ".join(
        f"{display}=o{internal_id}"
        for display, internal_id in sorted(option_mapping.items())
    )

def is_option_answer(value):

    return bool(
        re.fullmatch(
            r"[A-D,\s|]+",
            clean_option_text(value).upper()
        )
    )

def normalize_option_letters(value):

    text = clean_option_text(value).upper()

    if not is_option_answer(text):
        return text

    return "".join(
        sorted(
            re.findall(r"[A-D]", text)
        )
    )

def convert_display_answer_to_pdf_answer(
    answer,
    option_mapping
):

    answer = clean_option_text(answer)

    if not answer or not option_mapping or not is_option_answer(answer):
        return answer

    pdf_letters = []

    for display_option in re.findall(r"[A-D]", answer.upper()):

        internal_id = option_mapping.get(display_option, "")

        pdf_letter = internal_option_to_letter(internal_id)

        if pdf_letter:
            pdf_letters.append(pdf_letter)

    return "".join(
        sorted(pdf_letters)
    )

def convert_pdf_key_to_display_key(
    correct_key,
    option_mapping
):

    correct_key = clean_option_text(correct_key)

    if not correct_key or not option_mapping or not is_option_answer(correct_key):
        return correct_key

    pdf_to_display = {
        internal_option_to_letter(internal_id): display_option
        for display_option, internal_id in option_mapping.items()
    }

    display_parts = []

    for key_part in correct_key.upper().split("|"):

        display_letters = []

        for pdf_letter in re.findall(r"[A-D]", key_part):

            display_letter = pdf_to_display.get(pdf_letter, "")

            if display_letter:
                display_letters.append(display_letter)

        if display_letters:
            display_parts.append(
                "".join(
                    sorted(display_letters)
                )
            )

    return "|".join(display_parts)

def normalize_answer_for_compare(value):

    value = clean_option_text(value).upper()

    if not value:
        return ""

    option_value = re.sub(
        r"[\s,;]+",
        "",
        value
    )

    if re.fullmatch(r"[A-D]+", option_value):
        return "".join(
            sorted(option_value)
        )

    return re.sub(
        r"\s+",
        "",
        value
    )

def split_alternative_keys(correct_key):

    return [
        normalize_answer_for_compare(key)
        for key in str(correct_key).split("|")
        if normalize_answer_for_compare(key)
    ]

def is_m_answer_question(question_type):

    question_type = clean_text(question_type).lower()

    return (
        "msq" in question_type
        or "multiple select" in question_type
        or "multiple correct" in question_type
    )

def option_set(value):

    normalized = normalize_answer_for_compare(value)

    if re.fullmatch(r"[A-D]+", normalized):
        return set(normalized)

    return set()

def answers_match(candidate_answer, correct_key):

    candidate = normalize_answer_for_compare(
        candidate_answer
    )

    if not candidate:
        return False

    for correct in split_alternative_keys(correct_key):

        if candidate == correct:
            return True

        try:
            if Decimal(candidate) == Decimal(correct):
                return True

        except InvalidOperation:
            pass

    return False

def calculate_answer_score(
    answer,
    answer_key_record,
    question_type
):

    correct_key = answer_key_record.get("Key", "")
    positive_mark = answer_key_record.get("+mark", 0)
    negative_mark = answer_key_record.get("-ve mark", 0)

    if not clean_option_text(correct_key):
        return "Key Missing", 0

    if not clean_option_text(answer):
        return "Unattempted", 0

    if answers_match(answer, correct_key):
        return "Correct", positive_mark

    if is_m_answer_question(question_type):

        candidate_options = option_set(answer)

        correct_options = option_set(correct_key)

        if candidate_options and correct_options:

            wrong_options = candidate_options - correct_options

            if wrong_options:
                return "Wrong", negative_mark

            partial_marks = len(candidate_options)

            if partial_marks:
                return "Partial", partial_marks

    return "Wrong", negative_mark

# ============================================================
# SUMMARY HELPERS
# ============================================================

def normalize_paper_name(value):

    value = clean_text(value)

    if value.upper() == "P1":
        return "Paper 1"

    if value.upper() == "P2":
        return "Paper 2"

    return value

def normalize_subject_name(value):

    value = clean_text(value)

    if value.lower() == "math":
        return "Mathematics"

    return value

def calculate_candidate_summary(question_data):

    summary = {
        "total_marks": 0,
        "paper_marks": {
            "Paper 1": 0,
            "Paper 2": 0
        },
        "attempted": 0,
        "total_questions": len(question_data),
        "subject_paper_marks": {},
        "subject_result_counts": {}
    }

    for item in question_data:

        marks = item.get("Marks", 0)

        summary["total_marks"] += marks

        if clean_option_text(item.get("Answer", "")):
            summary["attempted"] += 1

        paper = normalize_paper_name(
            item.get("Key Paper", "")
            or item.get("Paper", "")
        )

        subject = normalize_subject_name(
            item.get("Key Subject", "")
        )

        if paper:
            summary["paper_marks"][paper] = (
                summary["paper_marks"].get(paper, 0)
                + marks
            )

        if subject and paper:
            subject_marks = summary["subject_paper_marks"].setdefault(
                subject,
                {
                    "Paper 1": 0,
                    "Paper 2": 0
                }
            )

            subject_marks[paper] = subject_marks.get(paper, 0) + marks

            result_counts = summary["subject_result_counts"].setdefault(
                subject,
                {
                    "Correct": 0,
                    "Partial": 0,
                    "Wrong": 0
                }
            )

            result = item.get("Result", "")

            if result in result_counts:
                result_counts[result] += 1

    return summary

def direct_cell_text(td):

    if not td:
        return ""

    return clean_text(
        "".join(
            td.find_all(
                string=True,
                recursive=False
            )
        )
    )

# ============================================================
# MENU VALUE EXTRACTION
# ============================================================

def extract_menu_value(menu_table, label):

    """
    Handles malformed HTML where td exists outside tr.
    Only matches the td's own text so nested wrapper cells do not
    accidentally match labels from child tables.
    """

    if not menu_table:
        return ""

    tds = menu_table.find_all("td")

    for idx in range(len(tds) - 1):

        key = direct_cell_text(tds[idx])

        value = clean_text(tds[idx + 1].get_text())

        if label.lower() in key.lower():
            return value

    return ""

def extract_given_answer(panel):

    question_row_table = panel.find(
        "table",
        class_="questionRowTbl"
    )

    return extract_menu_value(
        question_row_table,
        "Given Answer"
    )

# ============================================================
# QUESTION EXTRACTION
# ============================================================

def extract_questions_from_page(
    page,
    paper_name,
    answer_key_map
):

    print(f"Processing {paper_name}")

    results = []

    html = page.content()

    soup = BeautifulSoup(
        html,
        "html.parser"
    )

    # ========================================================
    # SECTION BLOCKS
    # ========================================================

    sections = soup.find_all(
        "div",
        class_="section-cntnr"
    )

    for section in sections:

        # ----------------------------------------------------
        # SECTION NAME
        # ----------------------------------------------------

        section_name = ""

        section_label = section.find(
            "div",
            class_="section-lbl"
        )

        if section_label:

            spans = section_label.find_all(
                "span"
            )

            if len(spans) >= 2:

                section_name = clean_text(
                    spans[1].get_text()
                )

        # ----------------------------------------------------
        # QUESTION PANELS
        # ----------------------------------------------------

        question_panels = section.find_all(
            "div",
            class_="question-pnl"
        )

        for panel in question_panels:

            try:

                # ==============================================
                # QUESTION NUMBER
                # ==============================================

                question_number = ""

                q_td = panel.find(
                    "td",
                    string=re.compile(r"Q\.\d+")
                )

                if q_td:

                    question_number = clean_text(
                        q_td.get_text()
                    )

                # ==============================================
                # MENU TABLE
                # ==============================================

                menu_table = panel.find(
                    "table",
                    class_="menu-tbl"
                )

                question_type = extract_menu_value(
                    menu_table,
                    "Question Type"
                )

                question_id = extract_menu_value(
                    menu_table,
                    "Question ID"
                )

                status = extract_menu_value(
                    menu_table,
                    "Status"
                )

               
                # ==============================================
                # OPTION MAPPING
                # ==============================================

                option_mapping = {}

                option_cells = panel.find_all("td")

                for td in option_cells:

                    text = clean_text(
                        td.get_text(" ", strip=True)
                    )

                    # Match only A. B. C. D.
                    match = re.match(
                        r"^([A-D])\.$",
                        text
                    )

                    if not match:
                        continue

                    display_option = match.group(1)

                    img = td.find("img")

                    internal_id = extract_internal_option_id(
                        img
                    )

                    if internal_id:

                        option_mapping[
                            display_option
                        ] = internal_id

                # ==============================================
                # ANSWER
                # ==============================================

                answer = ""

                chosen_option = extract_menu_value(
                    menu_table,
                    "Chosen Option"
                )

                chosen_option = clean_option_text(
                    chosen_option
                )

                given_answer = clean_option_text(
                    extract_given_answer(panel)
                )
                
                # MCQ / MSQ 
                if chosen_option:
                    answer = chosen_option 
                # SA / NAT 
                elif given_answer: 
                    answer = given_answer
                # ==============================================
                # INTERNAL ANSWER
                # ==============================================

                internal_answer = ""

                if answer and option_mapping:
                    internal_answer = convert_display_answer_to_pdf_answer(
                        answer,
                        option_mapping
                    )

                # ==============================================
                # CORRECT KEY
                # ==============================================

                answer_key_record = answer_key_map.get(
                    clean_text(question_id),
                    {}
                )

                correct_key = answer_key_record.get(
                    "Key",
                    ""
                )

                display_key = convert_pdf_key_to_display_key(
                    correct_key,
                    option_mapping
                )

                answer_for_scoring = internal_answer or answer

                result_status, marks_obtained = calculate_answer_score(
                    answer_for_scoring,
                    answer_key_record,
                    question_type
                )

                # ==============================================
                # RESULT OBJECT
                # ==============================================

                result = {

                    "Paper": answer_key_record.get("Paper", "") or paper_name,

                    "Subject": answer_key_record.get("Subject", ""),

                    "Section": answer_key_record.get("Sec", "") or section_name,

                    "Question": answer_key_record.get("Q", "") or question_number,

                    "Question Type": question_type,

                    "Question ID": question_id,

                    "Status": status,

                    "Answer": answer,

                    "Internal Answer": internal_answer,

                    "Display Key": display_key,

                    "Correct Key": correct_key,

                    "Result": result_status,

                    "Key Paper": answer_key_record.get("Paper", ""),

                    "Key Subject": answer_key_record.get("Subject", ""),

                    "+mark": answer_key_record.get("+mark", 0),

                    "-ve mark": answer_key_record.get("-ve mark", 0),

                    "Marks": marks_obtained,

                    "Option Map": option_mapping_to_text(option_mapping)
                }

                results.append(result)

            except Exception as e:

                print(
                    f"Question extraction failed: {e}"
                )

    print(
        f"Extracted {len(results)} questions"
    )

    return results

# ============================================================
# WRITE ALL CANDIDATES EXCEL SHEET
# ============================================================

ALL_STUDENTS_SHEET_NAME = "All Students"
ALL_STUDENTS_PAPER_ROW = 2
ALL_STUDENTS_QUESTION_ID_ROW = 3
ALL_STUDENTS_PDF_KEY_ROW = 4
ALL_STUDENTS_HEADER_ROW = 5
ALL_STUDENTS_DATA_START_ROW = 6
ALL_STUDENTS_SUMMARY_HEADERS = [
    "Total",
    "Paper 1",
    "Paper 2",
    "Correct",
    "Partial",
    "Wrong",
    "Chemistry Total",
    "Physics Total",
    "Mathematics Total"
]
ALL_STUDENTS_FIRST_QUESTION_COL = 3 + len(ALL_STUDENTS_SUMMARY_HEADERS)


def get_or_create_all_students_sheet(workbook):

    if ALL_STUDENTS_SHEET_NAME in workbook.sheetnames:
        ws = workbook[ALL_STUDENTS_SHEET_NAME]

    else:
        ws = workbook.create_sheet(
            title=ALL_STUDENTS_SHEET_NAME
        )

        ws.sheet_view.showGridLines = False

    sheet_index = workbook.index(ws)

    if sheet_index > 0:
        workbook.move_sheet(
            ws,
            offset=-sheet_index
        )

    return ws


def worksheet_link_target(sheet_name, cell_reference="A1"):

    escaped_sheet_name = str(sheet_name).replace("'", "''")

    return f"#'{escaped_sheet_name}'!{cell_reference}"


def question_identity(item):

    question_id = clean_text(
        item.get("Question ID", "")
    )

    if question_id:
        return question_id

    return "|".join(
        [
            clean_text(item.get("Paper", "")),
            clean_text(item.get("Subject", "")),
            clean_text(item.get("Section", "")),
            clean_text(item.get("Question", ""))
        ]
    )


def existing_question_columns(ws):

    question_columns = {}

    for col_num in range(
        ALL_STUDENTS_FIRST_QUESTION_COL,
        ws.max_column + 1
    ):

        question_id = clean_text(
            ws.cell(ALL_STUDENTS_QUESTION_ID_ROW, col_num).value
        )

        if question_id:
            question_columns[question_id] = col_num

    return question_columns


def find_candidate_row(ws, candidate_id):

    candidate_id = clean_text(candidate_id)

    if not candidate_id:
        return None

    for row_num in range(
        ALL_STUDENTS_DATA_START_ROW,
        ws.max_row + 1
    ):
        existing_id = clean_text(
            ws.cell(row_num, 1).value
        )

        if existing_id == candidate_id:
            return row_num

    return None


def candidate_exists_in_workbook(workbook, candidate_info):

    if ALL_STUDENTS_SHEET_NAME not in workbook.sheetnames:
        return False

    ws = workbook[ALL_STUDENTS_SHEET_NAME]

    return find_candidate_row(
        ws,
        candidate_info.get("Candidate ID", "")
    ) is not None


def refresh_paper_headers(ws, question_columns):

    for merged_range in list(ws.merged_cells.ranges):
        if (
            merged_range.min_row == ALL_STUDENTS_PAPER_ROW
            and merged_range.max_row == ALL_STUDENTS_PAPER_ROW
        ):
            ws.unmerge_cells(str(merged_range))

    last_paper_name = ""

    for col_num in range(
        ALL_STUDENTS_FIRST_QUESTION_COL,
        ws.max_column + 1
    ):

        paper = clean_text(
            ws.cell(ALL_STUDENTS_PAPER_ROW, col_num).value
        )

        if paper:
            last_paper_name = paper

        elif (
            last_paper_name
            and clean_text(
                ws.cell(ALL_STUDENTS_QUESTION_ID_ROW, col_num).value
            )
        ):
            ws.cell(
                ALL_STUDENTS_PAPER_ROW,
                col_num,
                last_paper_name
            )

    paper_ranges = []
    active_paper = None
    active_start = None
    last_col = None

    for col_num in range(
        ALL_STUDENTS_FIRST_QUESTION_COL,
        ws.max_column + 1
    ):

        question_id = clean_text(
            ws.cell(ALL_STUDENTS_QUESTION_ID_ROW, col_num).value
        )

        if not question_id:
            continue

        paper = clean_text(
            ws.cell(ALL_STUDENTS_PAPER_ROW, col_num).value
        )

        if paper != active_paper:
            if active_paper and active_start and last_col:
                paper_ranges.append(
                    (active_paper, active_start, last_col)
                )

            active_paper = paper
            active_start = col_num

        last_col = col_num

    if active_paper and active_start and last_col:
        paper_ranges.append(
            (active_paper, active_start, last_col)
        )

    paper_fill = PatternFill(
        "solid",
        fgColor="F7F2E8"
    )

    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    for paper, start_col, end_col in paper_ranges:

        if start_col != end_col:
            ws.merge_cells(
                start_row=ALL_STUDENTS_PAPER_ROW,
                start_column=start_col,
                end_row=ALL_STUDENTS_PAPER_ROW,
                end_column=end_col
            )

        cell = ws.cell(ALL_STUDENTS_PAPER_ROW, start_col)
        cell.value = paper
        cell.fill = paper_fill
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        for col_num in range(start_col, end_col + 1):
            ws.cell(ALL_STUDENTS_PAPER_ROW, col_num).border = thin_border


def ensure_all_students_layout(ws, question_data):

    dark_fill = PatternFill(
        "solid",
        fgColor="17324D"
    )

    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for row_num, label in (
        (ALL_STUDENTS_QUESTION_ID_ROW, "Question ID"),
        (ALL_STUDENTS_PDF_KEY_ROW, "PDF Key")
    ):
        ws.merge_cells(
            start_row=row_num,
            start_column=1,
            end_row=row_num,
            end_column=ALL_STUDENTS_FIRST_QUESTION_COL - 1
        )

        cell = ws.cell(row_num, 1, label)
        cell.fill = dark_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        for col_num in range(1, ALL_STUDENTS_FIRST_QUESTION_COL):
            ws.cell(row_num, col_num).border = thin_border

    fixed_headers = [
        "Candidate ID",
        "Candidate Name"
    ] + ALL_STUDENTS_SUMMARY_HEADERS

    for col_num, header in enumerate(fixed_headers, 1):
        cell = ws.cell(ALL_STUDENTS_HEADER_ROW, col_num, header)
        cell.font = (
            header_font
            if col_num >= 3
            else Font(bold=True)
        )
        if col_num >= 3:
            cell.fill = dark_fill
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    question_columns = existing_question_columns(ws)

    next_col = max(
        ws.max_column + 1,
        ALL_STUDENTS_FIRST_QUESTION_COL
    )

    for item in question_data:

        identity = question_identity(item)

        if not identity or identity in question_columns:
            continue

        col_num = next_col
        question_columns[identity] = col_num
        next_col += 1

        ws.cell(
            ALL_STUDENTS_PAPER_ROW,
            col_num,
            normalize_paper_name(
                item.get("Key Paper", "")
                or item.get("Paper", "")
            )
            or clean_text(item.get("Paper", ""))
        )
        ws.cell(ALL_STUDENTS_QUESTION_ID_ROW, col_num, identity)
        ws.cell(ALL_STUDENTS_PDF_KEY_ROW, col_num, item.get("Correct Key", ""))
        ws.cell(ALL_STUDENTS_HEADER_ROW, col_num, "Your PDF Ans")

        for row_num in range(
            ALL_STUDENTS_PAPER_ROW,
            ALL_STUDENTS_HEADER_ROW + 1
        ):
            cell = ws.cell(row_num, col_num)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            if row_num == ALL_STUDENTS_HEADER_ROW:
                cell.fill = dark_fill
                cell.font = header_font

    refresh_paper_headers(ws, question_columns)

    ws.freeze_panes = (
        f"{get_column_letter(ALL_STUDENTS_FIRST_QUESTION_COL)}"
        f"{ALL_STUDENTS_DATA_START_ROW}"
    )
    ws.row_dimensions[ALL_STUDENTS_PAPER_ROW].height = 20
    ws.row_dimensions[ALL_STUDENTS_QUESTION_ID_ROW].height = 20
    ws.row_dimensions[ALL_STUDENTS_PDF_KEY_ROW].height = 24
    ws.row_dimensions[ALL_STUDENTS_HEADER_ROW].height = 26

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 28

    for col_num in range(3, ALL_STUDENTS_FIRST_QUESTION_COL):
        ws.column_dimensions[get_column_letter(col_num)].width = 14

    for col_num in range(
        ALL_STUDENTS_FIRST_QUESTION_COL,
        ws.max_column + 1
    ):
        ws.column_dimensions[get_column_letter(col_num)].width = 14

    return question_columns


def write_candidate_sheet(
    workbook,
    candidate_info,
    question_data,
    detail_sheet_name=None
):

    ws = get_or_create_all_students_sheet(workbook)

    question_columns = ensure_all_students_layout(
        ws,
        question_data
    )

    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    correct_fill = PatternFill(
        "solid",
        fgColor="C6EFCE"
    )

    correct_font = Font(
        bold=True,
        color="006100"
    )

    summary = calculate_candidate_summary(
        question_data
    )

    result_counts = {
        "Correct": 0,
        "Partial": 0,
        "Wrong": 0
    }

    subject_totals = {
        "Chemistry": 0,
        "Physics": 0,
        "Mathematics": 0
    }

    for item in question_data:

        result = item.get("Result", "")

        if result in result_counts:
            result_counts[result] += 1

        subject = normalize_subject_name(
            item.get("Key Subject", "")
            or item.get("Subject", "")
        )

        if subject in subject_totals:
            subject_totals[subject] += item.get("Marks", 0)

    summary_values = [
        summary["total_marks"],
        summary["paper_marks"].get("Paper 1", 0),
        summary["paper_marks"].get("Paper 2", 0),
        result_counts["Correct"],
        result_counts["Partial"],
        result_counts["Wrong"],
        subject_totals["Chemistry"],
        subject_totals["Physics"],
        subject_totals["Mathematics"]
    ]

    existing_row = find_candidate_row(
        ws,
        candidate_info.get("Candidate ID", "")
    )

    if existing_row:
        return existing_row

    row_num = max(
        ws.max_row + 1,
        ALL_STUDENTS_DATA_START_ROW
    )

    candidate_id_cell = ws.cell(
        row_num,
        1,
        candidate_info.get("Candidate ID", "")
    )

    if detail_sheet_name:
        candidate_id_cell.hyperlink = worksheet_link_target(
            detail_sheet_name,
            "A1"
        )
        candidate_id_cell.font = Font(
            color="0563C1",
            underline="single"
        )

    ws.cell(
        row_num,
        2,
        candidate_info.get("Candidate Name", "")
    )

    for offset, value in enumerate(summary_values, 3):
        ws.cell(row_num, offset, value)

    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row_num, col_num)
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    for item in question_data:

        col_num = question_columns.get(
            question_identity(item)
        )

        if not col_num:
            continue

        answer_cell = ws.cell(
            row_num,
            col_num,
            item.get("Internal Answer", "")
            or item.get("Answer", "")
        )

        answer_cell.border = thin_border
        answer_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        if item.get("Result") == "Correct":
            answer_cell.fill = correct_fill
            answer_cell.font = correct_font

    if ws.max_row >= ALL_STUDENTS_DATA_START_ROW:
        ws.auto_filter.ref = (
            f"A{ALL_STUDENTS_HEADER_ROW}:"
            f"{get_column_letter(ws.max_column)}{ws.max_row}"
        )

    return row_num


# ============================================================
# WRITE DETAILED EXCEL SHEET
# ============================================================

def write_candidate_detail_sheet(
    workbook,
    candidate_info,
    question_data
):

    candidate_id = candidate_info.get(
        "Candidate ID",
        "Unknown"
    )

    sheet_name = str(candidate_id)[:31]

    ws = workbook.create_sheet(
        title=sheet_name
    )

    ws.sheet_view.showGridLines = False

    summary = calculate_candidate_summary(
        question_data
    )

    title_font = Font(
        bold=True,
        size=18,
        color="0B1F33"
    )

    label_font = Font(
        bold=True,
        size=10,
        color="42526E"
    )

    value_font = Font(
        bold=True,
        size=20,
        color="071E33"
    )

    table_header_fill = PatternFill(
        "solid",
        fgColor="E8EEF5"
    )

    dark_header_fill = PatternFill(
        "solid",
        fgColor="17324D"
    )

    card_fill = PatternFill(
        "solid",
        fgColor="F8FBFE"
    )

    section_fill = PatternFill(
        "solid",
        fgColor="EEF4FA"
    )

    zebra_fill = PatternFill(
        "solid",
        fgColor="FAFCFE"
    )

    thin_border = Border(
        left=Side(style="thin", color="D7DEE8"),
        right=Side(style="thin", color="D7DEE8"),
        top=Side(style="thin", color="D7DEE8"),
        bottom=Side(style="thin", color="D7DEE8")
    )

    result_styles = {
        "Correct": (
            PatternFill("solid", fgColor="E7F6EC"),
            Font(bold=True, color="17633A")
        ),
        "Partial": (
            PatternFill("solid", fgColor="FFF4D6"),
            Font(bold=True, color="8A5A00")
        ),
        "Wrong": (
            PatternFill("solid", fgColor="FDECEC"),
            Font(bold=True, color="A12626")
        ),
        "Unattempted": (
            PatternFill("solid", fgColor="F1F3F5"),
            Font(bold=True, color="5B6673")
        ),
        "Key Missing": (
            PatternFill("solid", fgColor="F3E8FF"),
            Font(bold=True, color="5E35B1")
        )
    }

    # ========================================================
    # SCORE SUMMARY
    # ========================================================

    ws.merge_cells("A1:H1")

    ws["A2"] = "Back to All Students"
    ws["A2"].hyperlink = worksheet_link_target(
        ALL_STUDENTS_SHEET_NAME,
        "A1"
    )
    ws["A2"].font = Font(
        bold=True,
        color="0563C1",
        underline="single"
    )

    candidate_name = clean_text(
        candidate_info.get("Candidate Name", "")
    )

    title_text = "JEE Advanced Scorecard"

    if candidate_name:
        title_text = f"{title_text} - {candidate_name}"

    ws["A1"] = title_text

    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws.row_dimensions[1].height = 30

    summary_cards = [
        ("Total Marks", summary["total_marks"]),
        ("Paper 1", summary["paper_marks"].get("Paper 1", 0)),
        ("Paper 2", summary["paper_marks"].get("Paper 2", 0)),
        (
            "Attempted",
            f'{summary["attempted"]}/{summary["total_questions"]}'
        )
    ]

    for index, (label, value) in enumerate(summary_cards):

        col = 1 + (index * 2)

        ws.cell(3, col, label)
        ws.cell(4, col, value)

        ws.cell(3, col).font = label_font
        ws.cell(4, col).font = value_font
        ws.cell(3, col).alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        ws.cell(4, col).alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        ws.merge_cells(
            start_row=3,
            start_column=col,
            end_row=3,
            end_column=col + 1
        )

        ws.merge_cells(
            start_row=4,
            start_column=col,
            end_row=4,
            end_column=col + 1
        )

        for row_num in (3, 4):

            for col_num in (col, col + 1):

                cell = ws.cell(row_num, col_num)
                cell.fill = card_fill
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 32

    subject_summary_row = 7

    ws.cell(subject_summary_row, 1, "Subject Summary")
    ws.cell(subject_summary_row, 1).font = title_font

    subject_summary_headers = [
        "Subject",
        "Paper 1",
        "Paper 2",
        "Total",
        "Correct",
        "Partial",
        "Wrong"
    ]

    subject_header_row = subject_summary_row + 2

    for col_num, header in enumerate(
        subject_summary_headers,
        1
    ):

        cell = ws.cell(subject_header_row, col_num, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = dark_header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[subject_header_row].height = 24

    subject_order = [
        "Mathematics",
        "Physics",
        "Chemistry"
    ]

    for offset, subject in enumerate(
        subject_order,
        1
    ):

        current_row = subject_header_row + offset

        subject_marks = summary["subject_paper_marks"].get(
            subject,
            {}
        )

        result_counts = summary["subject_result_counts"].get(
            subject,
            {}
        )

        paper_1_marks = subject_marks.get("Paper 1", 0)
        paper_2_marks = subject_marks.get("Paper 2", 0)

        values = [
            subject.replace("Mathematics", "Math"),
            paper_1_marks,
            paper_2_marks,
            paper_1_marks + paper_2_marks,
            result_counts.get("Correct", 0),
            result_counts.get("Partial", 0),
            result_counts.get("Wrong", 0)
        ]

        for col_num, value in enumerate(values, 1):

            cell = ws.cell(current_row, col_num, value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if offset % 2 == 0:
                cell.fill = zebra_fill

            if col_num == 4:
                cell.font = Font(bold=True, color="0B1F33")

            if col_num == 5:
                cell.font = Font(bold=True, color="17633A")

            if col_num == 6:
                cell.font = Font(bold=True, color="8A5A00")

            if col_num == 7:
                cell.font = Font(bold=True, color="A12626")

    summary_row = subject_header_row + len(subject_order) + 3

    ws.cell(summary_row, 3, "Breakdown")
    ws.cell(summary_row, 4, "Marks")
    ws.cell(summary_row, 3).font = Font(bold=True, color="FFFFFF")
    ws.cell(summary_row, 4).font = Font(bold=True, color="FFFFFF")

    for col_num in (3, 4):

        ws.cell(summary_row, col_num).fill = dark_header_fill
        ws.cell(summary_row, col_num).border = thin_border
        ws.cell(summary_row, col_num).alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    breakdown_subject_order = [
        "Chemistry",
        "Physics",
        "Mathematics"
    ]

    summary_rows = []

    for subject in breakdown_subject_order:

        subject_marks = summary["subject_paper_marks"].get(
            subject,
            {}
        )

        paper_1_marks = subject_marks.get("Paper 1", 0)
        paper_2_marks = subject_marks.get("Paper 2", 0)

        summary_rows.extend(
            [
                (f"{subject} 1", paper_1_marks, False),
                (f"{subject} 2", paper_2_marks, False),
                (
                    f"{subject} Total",
                    paper_1_marks + paper_2_marks,
                    True
                )
            ]
        )

    summary_rows.extend(
        [
            ("Total", summary["total_marks"], True),
            (
                "Positive",
                sum(
                    item.get("Marks", 0)
                    for item in question_data
                    if item.get("Marks", 0) > 0
                ),
                True
            )
        ]
    )

    for offset, (label, value, is_total) in enumerate(
        summary_rows,
        1
    ):

        current_row = summary_row + offset

        ws.cell(current_row, 3, label)
        ws.cell(current_row, 4, value)

        for col_num in (3, 4):

            cell = ws.cell(current_row, col_num)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            if is_total:
                cell.font = Font(bold=True)
                cell.fill = section_fill

            elif offset % 2 == 0:
                cell.fill = zebra_fill

    # ========================================================
    # CANDIDATE DETAILS
    # ========================================================

    row = summary_row + len(summary_rows) + 3

    ws[f"A{row}"] = "Candidate Details"

    ws[f"A{row}"].font = title_font

    row += 2

    for key, value in candidate_info.items():

        ws[f"A{row}"] = key

        ws[f"A{row}"].font = Font(
            bold=True,
            color="42526E"
        )

        ws[f"B{row}"] = value
        ws[f"A{row}"].fill = section_fill
        ws[f"A{row}"].border = thin_border
        ws[f"B{row}"].border = thin_border

        row += 1

    # ========================================================
    # TABLE HEADERS
    # ========================================================

    start_row = row + 2

    headers = [

        "Paper",
        "Subject",
        "Sec",
        "Q",
        "Question ID",
        "Option IDs",

        "Your Ans",
        "Your PDF Ans",
        "Display Key",
        "PDF Key",
        "Result",
        "Marks",
        "Question Type",
        "Status",
        "+mark",
        "-ve mark"
    ]

    for col_num, header in enumerate(
        headers,
        1
    ):

        cell = ws.cell(
            row=start_row,
            column=col_num
        )

        cell.value = header

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = dark_header_fill
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    ws.row_dimensions[start_row].height = 28

    # ========================================================
    # DATA ROWS
    # ========================================================

    for idx, item in enumerate(
        question_data,
        start_row + 1
    ):

        ws.cell(idx, 1, item["Paper"])
        ws.cell(idx, 2, item["Subject"])
        ws.cell(idx, 3, item["Section"])
        ws.cell(idx, 4, item["Question"])
        ws.cell(idx, 5, item["Question ID"])
        ws.cell(idx, 6, item["Option Map"])

        ws.cell(idx, 7, item["Answer"])
        ws.cell(idx, 8, item["Internal Answer"])
        ws.cell(idx, 9, item["Display Key"])
        ws.cell(idx, 10, item["Correct Key"])
        ws.cell(idx, 11, item["Result"])
        ws.cell(idx, 12, item["Marks"])
        ws.cell(idx, 13, item["Question Type"])
        ws.cell(idx, 14, item["Status"])
        ws.cell(idx, 15, item["+mark"])
        ws.cell(idx, 16, item["-ve mark"])

        for col_num in range(1, len(headers) + 1):

            cell = ws.cell(idx, col_num)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            if (idx - start_row) % 2 == 0:
                cell.fill = zebra_fill

        result_cell = ws.cell(idx, 11)
        result_fill, result_font = result_styles.get(
            item["Result"],
            (
                PatternFill("solid", fgColor="FFFFFF"),
                Font(color="000000")
            )
        )
        result_cell.fill = result_fill
        result_cell.font = result_font

        marks_cell = ws.cell(idx, 12)
        marks_cell.font = Font(
            bold=True,
            color=(
                "17633A"
                if item["Marks"] > 0
                else "A12626"
                if item["Marks"] < 0
                else "5B6673"
            )
        )

    if question_data:
        last_data_row = start_row + len(question_data)
        last_data_col = len(headers)
        ws.auto_filter.ref = (
            f"A{start_row}:"
            f"{get_column_letter(last_data_col)}{last_data_row}"
        )

    # ========================================================
    # AUTO WIDTH
    # ========================================================

    for column_index, column_cells in enumerate(
        ws.columns,
        1
    ):

        length = 0

        column = get_column_letter(column_index)

        for cell in column_cells:

            try:

                if len(str(cell.value)) > length:

                    length = len(str(cell.value))

            except:
                pass

        ws.column_dimensions[column].width = min(
            length + 5,
            80
        )

    return ws.title

# ============================================================
# PROCESS SINGLE CANDIDATE
# ============================================================

def process_candidate(
    context,
    candidate,
    answer_key_map
):

    username = clean_text(
        candidate["USERNAME"]
    )

    password = clean_text(
        candidate["PASSWORD"]
    )

    dob = format_dob(
        candidate["DOB"]
    )

    print(
        f"\nProcessing Candidate: {username}"
    )

    page = context.new_page()

    all_results = []

    candidate_info = {}

    try:

        # ====================================================
        # LOGIN
        # ====================================================

        page.goto(
            LOGIN_URL,
            wait_until="networkidle"
        )

        page.fill(
            'input[name="AdvAppNo"]',
            username
        )

        page.fill(
            'input[name="mobileNo"]',
            password
        )

        page.fill(
            'input[name="datepicker"]',
            dob
        )

        page.click(
            'button[type="submit"]'
        )

        page.wait_for_load_state(
            "networkidle"
        )

        # ====================================================
        # OPEN RESPONSE PAGE
        # ====================================================

        try:

            page.click(
                "text=view Response",
                timeout=5000
            )

        except:

            page.goto(
                DISPLAY_RESPONSE_URL,
                wait_until="networkidle"
            )

        page.wait_for_load_state(
            "networkidle"
        )

        # ====================================================
        # FIND PAPER LINKS
        # ====================================================

        links = page.locator(
            'a:has-text("View Here")'
        )

        count = links.count()

        print(f"Found {count} papers")

        if count == 0:

            raise Exception(
                "No paper links found"
            )

        # ====================================================
        # PROCESS PAPERS
        # ====================================================

        for i in range(count):

            try:

                link = links.nth(i)

                href = link.get_attribute(
                    "href"
                )

                if not href:
                    continue

                full_url = urljoin(
                    page.url,
                    href
                )

                paper_name = f"Paper {i+1}"

                print(
                    f"Opening {paper_name}"
                )

                paper_page = context.new_page()

                paper_page.goto(
                    full_url,
                    wait_until="networkidle"
                )

                # Candidate Info
                if not candidate_info:

                    candidate_info = (
                        extract_candidate_info(
                            paper_page
                        )
                    )

                # Extract Questions
                paper_results = (
                    extract_questions_from_page(
                        paper_page,
                        paper_name,
                        answer_key_map
                    )
                )

                all_results.extend(
                    paper_results
                )

                paper_page.close()

            except Exception as e:

                print(
                    f"Paper processing failed: {e}"
                )

        page.close()

        return candidate_info, all_results, ""

    except Exception as e:

        print(
            f"Candidate failed: {e}"
        )

        traceback.print_exc()

        page.close()

        return None, None, str(e)

# ============================================================
# FAILED CANDIDATE LOG
# ============================================================

def build_failed_candidate_record(
    candidate,
    status,
    error_message
):

    return {
        "USERNAME": clean_text(
            candidate.get("USERNAME", "")
        ),
        "PASSWORD": clean_text(
            candidate.get("PASSWORD", "")
        ),
        "DOB": clean_text(
            candidate.get("DOB", "")
        ),
        "Status": status,
        "Error": clean_text(
            error_message
        )
    }

def write_failed_candidates_excel(
    failed_candidates
):

    if not failed_candidates:
        return

    workbook = Workbook()

    ws = workbook.active

    ws.title = "Failed Candidates"
    ws.sheet_view.showGridLines = False

    headers = [
        "USERNAME",
        "PASSWORD",
        "DOB",
        "Status",
        "Error"
    ]

    for col_num, header in enumerate(
        headers,
        1
    ):

        cell = ws.cell(
            row=1,
            column=col_num
        )

        cell.value = header

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )
        cell.fill = PatternFill("solid", fgColor="17324D")
        cell.border = Border(
            left=Side(style="thin", color="D7DEE8"),
            right=Side(style="thin", color="D7DEE8"),
            top=Side(style="thin", color="D7DEE8"),
            bottom=Side(style="thin", color="D7DEE8")
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    ws.row_dimensions[1].height = 26

    for row_num, candidate in enumerate(
        failed_candidates,
        2
    ):

        for col_num, header in enumerate(
            headers,
            1
        ):

            cell = ws.cell(
                row=row_num,
                column=col_num,
                value=candidate.get(header, "")
            )
            cell.border = Border(
                left=Side(style="thin", color="D7DEE8"),
                right=Side(style="thin", color="D7DEE8"),
                top=Side(style="thin", color="D7DEE8"),
                bottom=Side(style="thin", color="D7DEE8")
            )
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

            if row_num % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="FAFCFE")

    if failed_candidates:
        ws.auto_filter.ref = f"A1:E{len(failed_candidates) + 1}"
        ws.freeze_panes = "A2"

    for column_cells in ws.columns:

        length = 0

        column = column_cells[0].column_letter

        for cell in column_cells:

            try:

                length = max(
                    length,
                    len(str(cell.value))
                )

            except:
                pass

        ws.column_dimensions[column].width = min(
            length + 5,
            80
        )

    workbook.save(FAILED_OUTPUT_FILE)

# ============================================================
# MAIN
# ============================================================

def main():

    candidates = read_candidate_excel()

    answer_key_map = read_answer_key()

    output_mode = ask_output_mode()

    workbook, appended_to_existing = create_output_workbook(output_mode)

    if appended_to_existing:
        print(
            f"\nAppending responses to existing file:\n{OUTPUT_FILE}"
        )

    else:
        print(
            f"\nCreating a new output file:\n{OUTPUT_FILE}"
        )

    failed_candidates = []

    with sync_playwright() as p:

        browser = p.chromium.launch(
            headless=HEADLESS,
            slow_mo=SLOW_MO
        )

        for candidate in candidates:

            context = browser.new_context()

            try:

                candidate_info, results, error_message = (
                    process_candidate(
                        context,
                        candidate,
                        answer_key_map
                    )
                )

                if candidate_info and results:

                    candidate_id = clean_text(
                        candidate_info.get("Candidate ID", "")
                    )

                    if candidate_exists_in_workbook(
                        workbook,
                        candidate_info
                    ):
                        print(
                            f"Skipping {candidate_id}: candidate already exists in output file"
                        )
                        continue

                    detail_sheet_name = write_candidate_detail_sheet(
                        workbook,
                        candidate_info,
                        results
                    )

                    write_candidate_sheet(
                        workbook,
                        candidate_info,
                        results,
                        detail_sheet_name
                    )

                else:

                    failed_candidates.append(
                        build_failed_candidate_record(
                            candidate,
                            "Failed",
                            error_message or "No candidate details / question data extracted"
                        )
                    )

            except Exception as e:

                print(
                    f"Processing failed: {e}"
                )

                failed_candidates.append(
                    build_failed_candidate_record(
                        candidate,
                        "Failed",
                        str(e)
                    )
                )

            finally:

                context.close()

        browser.close()

    if not workbook.sheetnames:

        ws = workbook.create_sheet(
            title="No Successful Candidates"
        )

        ws["A1"] = "No successful candidates processed"

        ws["A1"].font = Font(
            bold=True
        )

    workbook.save(OUTPUT_FILE)

    write_failed_candidates_excel(
        failed_candidates
    )

    print(
        f"\nFinal Output Saved:\n{OUTPUT_FILE}"
    )

    if failed_candidates:

        print(
            f"Failed Candidates Saved:\n{FAILED_OUTPUT_FILE}"
        )

# ============================================================
# ENTRY
# ============================================================

if __name__ == "__main__":

    main()
